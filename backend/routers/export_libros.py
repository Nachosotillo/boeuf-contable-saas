"""
routers/export_libros.py — Libros legales descargables en XLSX (formato venezolano).

Genera, en vivo y con encabezado legal (razón social, RIF, período), los libros que
exige la normativa venezolana, listos para imprimir/mostrar:

  GET /libros/diario.xlsx?mes=
  GET /libros/mayor.xlsx?mes=&cuenta=
  GET /libros/compras.xlsx?mes=        (Providencia SNAT/2003/1.677)
  GET /libros/ventas.xlsx?mes=         (Providencia SNAT/2003/1.677)
  GET /libros/inventarios.xlsx?articulo=   (Libro de Inventarios / kardéx)
  GET /libros/catalogo.xlsx
  GET /libros/balance.xlsx?mes=
  GET /libros/todos.xlsx?mes=          (un solo libro con todas las hojas)

Requiere openpyxl en requirements.txt.
"""
import io
from decimal import Decimal
from datetime import date
from collections import defaultdict

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, extract

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db
from models import (Empresa, Asiento, LineaAsiento, CatalogoCuenta, Usuario,
                    LibroIvaCompra, LibroIvaVenta, ArticuloInventario, MovimientoInventario)
from routers.auth import get_current_user

router = APIRouter()

FUENTE = "Arial"
F_TITULO = Font(name=FUENTE, bold=True, size=14)
F_SUB    = Font(name=FUENTE, bold=True, size=10)
F_HEAD   = Font(name=FUENTE, bold=True, size=9, color="FFFFFF")
F_CELL   = Font(name=FUENTE, size=9)
F_TOTAL  = Font(name=FUENTE, bold=True, size=9)
FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
FILL_TOT  = PatternFill("solid", fgColor="D9E1F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right")
LEFT   = Alignment(horizontal="left")
THIN = Side(style="thin", color="999999")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = '#,##0.00;(#,##0.00);"-"'
MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _periodo(mes):
    return f"{MESES[mes]} 2025" if mes else "Ejercicio 2025 (acumulado)"


def _encabezado(ws, empresa: Empresa, titulo: str, periodo: str, ncols: int):
    last = get_column_letter(ncols)
    def linea(row, texto, font, align=CENTER):
        ws.merge_cells(f"A{row}:{last}{row}")
        c = ws[f"A{row}"]; c.value = texto; c.font = font; c.alignment = align
    linea(1, empresa.nombre_razon_social.upper(), F_TITULO)
    linea(2, f"RIF: {empresa.rif}", F_SUB)
    if empresa.direccion:
        linea(3, empresa.direccion, F_CELL)
    linea(4, titulo.upper(), F_SUB)
    linea(5, f"Período: {periodo}   ·   Expresado en Bolívares (Bs.)", F_CELL)
    return 7  # primera fila de la tabla


def _tabla_header(ws, row, columnas):
    for j, (titulo, _w) in enumerate(columnas, 1):
        c = ws.cell(row, j, titulo)
        c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = CENTER; c.border = BORDE
        ws.column_dimensions[get_column_letter(j)].width = _w
    ws.freeze_panes = ws.cell(row + 1, 1)
    return row + 1


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


async def _empresa(db, empresa_id) -> Empresa:
    e = (await db.execute(select(Empresa).where(Empresa.id == empresa_id))).scalar_one_or_none()
    if not e:
        raise HTTPException(404, "Empresa no encontrada")
    return e


async def _lineas(db, empresa_id, mes):
    q = (select(LineaAsiento, Asiento, CatalogoCuenta)
         .join(Asiento, LineaAsiento.asiento_id == Asiento.id)
         .join(CatalogoCuenta, LineaAsiento.cuenta_id == CatalogoCuenta.id)
         .where(Asiento.empresa_id == empresa_id, Asiento.es_prueba == False)
         .order_by(Asiento.fecha, Asiento.numero_asiento, LineaAsiento.id))
    if mes:
        q = q.where(Asiento.mes == mes)
    return (await db.execute(q)).all()


def _set(ws, r, c, v, *, font=F_CELL, align=LEFT, num=False, border=True):
    cell = ws.cell(r, c, v)
    cell.font = font; cell.alignment = (RIGHT if num else align)
    if num:
        cell.number_format = NUM
    if border:
        cell.border = BORDE
    return cell


# ─── LIBRO DIARIO ────────────────────────────────────────────────────────────
def _hoja_diario(ws, empresa, filas, periodo):
    cols = [("Fecha", 12), ("N° Asiento", 12), ("Código", 12), ("Cuenta", 38),
            ("Descripción", 40), ("Débito (Bs.)", 16), ("Crédito (Bs.)", 16)]
    r = _encabezado(ws, empresa, "Libro Diario", periodo, len(cols))
    r = _tabla_header(ws, r, cols)
    ini = r; td = th = Decimal("0")
    asiento_prev = None
    for ln, a, c in filas:
        primera = a.numero_asiento != asiento_prev
        asiento_prev = a.numero_asiento
        _set(ws, r, 1, a.fecha.isoformat() if primera else "")
        _set(ws, r, 2, a.numero_asiento if primera else "")
        _set(ws, r, 3, c.codigo)
        _set(ws, r, 4, c.nombre, align=(LEFT if ln.debe else Alignment(horizontal="left", indent=2)))
        _set(ws, r, 5, ln.descripcion or (a.descripcion if primera else ""))
        _set(ws, r, 6, float(ln.debe) if ln.debe else None, num=True)
        _set(ws, r, 7, float(ln.haber) if ln.haber else None, num=True)
        td += ln.debe; th += ln.haber; r += 1
    _set(ws, r, 1, "TOTALES", font=F_TOTAL); 
    for j in (2, 3, 4, 5):
        ws.cell(r, j).fill = FILL_TOT; ws.cell(r, j).border = BORDE
    ws.cell(r, 1).fill = FILL_TOT
    t6 = _set(ws, r, 6, None, font=F_TOTAL, num=True); t6.value = f"=SUM(F{ini}:F{r-1})"; t6.fill = FILL_TOT
    t7 = _set(ws, r, 7, None, font=F_TOTAL, num=True); t7.value = f"=SUM(G{ini}:G{r-1})"; t7.fill = FILL_TOT


# ─── LIBRO MAYOR ─────────────────────────────────────────────────────────────
def _hoja_mayor(ws, empresa, filas, periodo, cuenta_filtro):
    cols = [("Código", 12), ("Cuenta", 38), ("Fecha", 12), ("N° Asiento", 12),
            ("Descripción", 38), ("Débito", 15), ("Crédito", 15), ("Saldo", 16)]
    r = _encabezado(ws, empresa, "Libro Mayor", periodo, len(cols))
    r = _tabla_header(ws, r, cols)
    porc = defaultdict(list)
    for ln, a, c in filas:
        if cuenta_filtro and c.codigo != cuenta_filtro:
            continue
        porc[(c.codigo, c.nombre)].append((a, ln))
    for (cod, nom) in sorted(porc.keys()):
        saldo = Decimal("0")
        for a, ln in porc[(cod, nom)]:
            saldo += ln.debe - ln.haber
            _set(ws, r, 1, cod); _set(ws, r, 2, nom); _set(ws, r, 3, a.fecha.isoformat())
            _set(ws, r, 4, a.numero_asiento); _set(ws, r, 5, ln.descripcion or a.descripcion or "")
            _set(ws, r, 6, float(ln.debe) if ln.debe else None, num=True)
            _set(ws, r, 7, float(ln.haber) if ln.haber else None, num=True)
            _set(ws, r, 8, float(saldo), num=True); r += 1
        _set(ws, r, 5, f"Saldo final {cod}", font=F_TOTAL)
        for j in (1, 2, 3, 4, 6, 7): ws.cell(r, j).fill = FILL_TOT; ws.cell(r, j).border = BORDE
        ws.cell(r, 5).fill = FILL_TOT
        s = _set(ws, r, 8, float(saldo), font=F_TOTAL, num=True); s.fill = FILL_TOT
        r += 1


# ─── LIBROS IVA (SENIAT) ─────────────────────────────────────────────────────
def _hoja_compras(ws, empresa, regs, periodo):
    cols = [("Op. N°", 8), ("Fecha", 12), ("RIF Proveedor", 16), ("Razón Social", 34),
            ("N° Factura", 14), ("Total c/IVA", 16), ("Base Imponible", 16),
            ("Alíc.", 8), ("IVA Crédito Fiscal", 16), ("Retención IVA", 14), ("IGTF", 12)]
    r = _encabezado(ws, empresa, "Libro de Compras (IVA)", periodo, len(cols))
    r = _tabla_header(ws, r, cols); ini = r
    for i, x in enumerate(regs, 1):
        _set(ws, r, 1, i, align=CENTER); _set(ws, r, 2, x.fecha.isoformat())
        _set(ws, r, 3, x.rif_proveedor); _set(ws, r, 4, x.proveedor); _set(ws, r, 5, x.numero_factura)
        _set(ws, r, 6, float(x.total_factura), num=True); _set(ws, r, 7, float(x.base_imponible), num=True)
        _set(ws, r, 8, f"{float(x.alicuota_iva)*100:.0f}%", align=CENTER)
        _set(ws, r, 9, float(x.iva_credito_fiscal), num=True)
        _set(ws, r, 10, float(x.retencion_iva_75), num=True); _set(ws, r, 11, float(x.igtf_aplicado), num=True); r += 1
    _set(ws, r, 5, "TOTALES", font=F_TOTAL); ws.cell(r, 5).fill = FILL_TOT
    for col in (6, 7, 9, 10, 11):
        L = get_column_letter(col)
        t = _set(ws, r, col, None, font=F_TOTAL, num=True); t.value = f"=SUM({L}{ini}:{L}{r-1})"; t.fill = FILL_TOT
    for col in (1, 2, 3, 4, 8):
        ws.cell(r, col).fill = FILL_TOT; ws.cell(r, col).border = BORDE


def _hoja_ventas(ws, empresa, regs, periodo):
    cols = [("Op. N°", 8), ("Fecha", 12), ("RIF Cliente", 16), ("Razón Social", 34),
            ("N° Factura", 14), ("Total c/IVA", 16), ("Base Imponible", 16),
            ("Alíc.", 8), ("IVA Débito Fiscal", 16), ("IVA Retenido", 14)]
    r = _encabezado(ws, empresa, "Libro de Ventas (IVA)", periodo, len(cols))
    r = _tabla_header(ws, r, cols); ini = r
    for i, x in enumerate(regs, 1):
        _set(ws, r, 1, i, align=CENTER); _set(ws, r, 2, x.fecha.isoformat())
        _set(ws, r, 3, x.rif_cliente); _set(ws, r, 4, x.cliente); _set(ws, r, 5, x.numero_factura)
        _set(ws, r, 6, float(x.total_factura), num=True); _set(ws, r, 7, float(x.base_imponible), num=True)
        _set(ws, r, 8, f"{float(x.alicuota_iva)*100:.0f}%", align=CENTER)
        _set(ws, r, 9, float(x.iva_debito_fiscal), num=True)
        _set(ws, r, 10, float(x.retencion_iva_recibida), num=True); r += 1
    _set(ws, r, 5, "TOTALES", font=F_TOTAL); ws.cell(r, 5).fill = FILL_TOT
    for col in (6, 7, 9, 10):
        L = get_column_letter(col)
        t = _set(ws, r, col, None, font=F_TOTAL, num=True); t.value = f"=SUM({L}{ini}:{L}{r-1})"; t.fill = FILL_TOT
    for col in (1, 2, 3, 4, 8):
        ws.cell(r, col).fill = FILL_TOT; ws.cell(r, col).border = BORDE


# ─── LIBRO DE INVENTARIOS (kardéx) ───────────────────────────────────────────
def _hoja_inventarios(ws, empresa, articulos, movs_por_art, periodo):
    cols = [("Artículo", 30), ("Fecha", 12), ("Concepto", 30), ("Tipo", 10),
            ("Cant.", 12), ("Costo Unit.", 14), ("Costo Total", 16),
            ("Saldo Uds.", 12), ("Saldo Valor", 16)]
    r = _encabezado(ws, empresa, "Libro de Inventarios (Kárdex — Promedio Ponderado)", periodo, len(cols))
    r = _tabla_header(ws, r, cols)
    for art in articulos:
        for m in movs_por_art.get(art.id, []):
            _set(ws, r, 1, f"{art.codigo_sku} — {art.descripcion}")
            _set(ws, r, 2, m.fecha.isoformat()); _set(ws, r, 3, m.descripcion)
            _set(ws, r, 4, m.tipo.value if m.tipo else "", align=CENTER)
            _set(ws, r, 5, float(m.cantidad), num=True); _set(ws, r, 6, float(m.costo_unitario), num=True)
            _set(ws, r, 7, float(m.costo_total), num=True); _set(ws, r, 8, float(m.saldo_unidades), num=True)
            _set(ws, r, 9, float(getattr(m, "saldo_valor", 0) or 0), num=True); r += 1


# ─── CATÁLOGO ────────────────────────────────────────────────────────────────
def _hoja_catalogo(ws, empresa, cuentas):
    cols = [("Código", 14), ("Denominación de la Cuenta", 44), ("Tipo", 12),
            ("Naturaleza", 12), ("Estado Financiero", 22), ("Subcategoría", 22)]
    r = _encabezado(ws, empresa, "Catálogo / Plan de Cuentas", _periodo(None), len(cols))
    r = _tabla_header(ws, r, cols)
    for c in cuentas:
        sangria = (c.codigo.count(".") if c.tipo and c.tipo.value != "Cuenta" else c.codigo.count("."))
        _set(ws, r, 1, c.codigo)
        nom = _set(ws, r, 2, c.nombre)
        nom.alignment = Alignment(horizontal="left", indent=min(c.codigo.count("."), 5))
        if c.tipo and c.tipo.value != "Cuenta":
            nom.font = F_SUB
        _set(ws, r, 3, c.tipo.value if c.tipo else "", align=CENTER)
        _set(ws, r, 4, c.naturaleza.value if c.naturaleza else "", align=CENTER)
        _set(ws, r, 5, c.estado_financiero.value if c.estado_financiero else "")
        _set(ws, r, 6, c.subcategoria or ""); r += 1


# ─── BALANCE DE COMPROBACIÓN ─────────────────────────────────────────────────
def _hoja_balance(ws, empresa, filas, periodo):
    cols = [("Código", 14), ("Cuenta", 40), ("Débitos", 16), ("Créditos", 16),
            ("Saldo Deudor", 16), ("Saldo Acreedor", 16)]
    r = _encabezado(ws, empresa, "Balance de Comprobación", periodo, len(cols))
    r = _tabla_header(ws, r, cols); ini = r
    acum = defaultdict(lambda: {"nom": "", "d": Decimal("0"), "h": Decimal("0")})
    for ln, a, c in filas:
        s = acum[c.codigo]; s["nom"] = c.nombre; s["d"] += ln.debe; s["h"] += ln.haber
    for cod in sorted(acum):
        s = acum[cod]; deu = s["d"]-s["h"] if s["d"]>s["h"] else Decimal("0"); acr = s["h"]-s["d"] if s["h"]>s["d"] else Decimal("0")
        _set(ws, r, 1, cod); _set(ws, r, 2, s["nom"])
        _set(ws, r, 3, float(s["d"]), num=True); _set(ws, r, 4, float(s["h"]), num=True)
        _set(ws, r, 5, float(deu), num=True); _set(ws, r, 6, float(acr), num=True); r += 1
    _set(ws, r, 2, "TOTALES", font=F_TOTAL); ws.cell(r, 2).fill = FILL_TOT; ws.cell(r, 1).fill = FILL_TOT; ws.cell(r,1).border=BORDE
    for col in (3, 4, 5, 6):
        L = get_column_letter(col)
        t = _set(ws, r, col, None, font=F_TOTAL, num=True); t.value = f"=SUM({L}{ini}:{L}{r-1})"; t.fill = FILL_TOT


# ─── Endpoints ───────────────────────────────────────────────────────────────
@router.get("/diario.xlsx")
async def libro_diario(mes: int | None = Query(None, ge=1, le=12),
                       current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    filas = await _lineas(db, emp.id, mes)
    wb = Workbook(); _hoja_diario(wb.active, emp, filas, _periodo(mes)); wb.active.title = "Libro Diario"
    return _xlsx_response(wb, f"libro_diario_{mes or 'Q1'}.xlsx")


@router.get("/mayor.xlsx")
async def libro_mayor(mes: int | None = Query(None, ge=1, le=12), cuenta: str | None = Query(None),
                      current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    filas = await _lineas(db, emp.id, mes)
    wb = Workbook(); _hoja_mayor(wb.active, emp, filas, _periodo(mes), cuenta); wb.active.title = "Libro Mayor"
    return _xlsx_response(wb, f"libro_mayor_{mes or 'Q1'}.xlsx")


@router.get("/compras.xlsx")
async def libro_compras(mes: int | None = Query(None, ge=1, le=12),
                        current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    q = select(LibroIvaCompra).where(LibroIvaCompra.empresa_id == emp.id)
    if mes: q = q.where(extract("month", LibroIvaCompra.fecha) == mes)
    regs = (await db.execute(q.order_by(LibroIvaCompra.fecha, LibroIvaCompra.id))).scalars().all()
    wb = Workbook(); _hoja_compras(wb.active, emp, regs, _periodo(mes)); wb.active.title = "Libro de Compras"
    return _xlsx_response(wb, f"libro_compras_{mes or 'Q1'}.xlsx")


@router.get("/ventas.xlsx")
async def libro_ventas(mes: int | None = Query(None, ge=1, le=12),
                       current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    q = select(LibroIvaVenta).where(LibroIvaVenta.empresa_id == emp.id)
    if mes: q = q.where(extract("month", LibroIvaVenta.fecha) == mes)
    regs = (await db.execute(q.order_by(LibroIvaVenta.fecha, LibroIvaVenta.id))).scalars().all()
    wb = Workbook(); _hoja_ventas(wb.active, emp, regs, _periodo(mes)); wb.active.title = "Libro de Ventas"
    return _xlsx_response(wb, f"libro_ventas_{mes or 'Q1'}.xlsx")


@router.get("/inventarios.xlsx")
async def libro_inventarios(current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    arts = (await db.execute(select(ArticuloInventario).where(
        ArticuloInventario.empresa_id == emp.id).order_by(ArticuloInventario.codigo_sku))).scalars().all()
    movs = (await db.execute(select(MovimientoInventario).where(
        MovimientoInventario.empresa_id == emp.id).order_by(MovimientoInventario.fecha, MovimientoInventario.id))).scalars().all()
    por_art = defaultdict(list)
    for m in movs: por_art[m.articulo_id].append(m)
    wb = Workbook(); _hoja_inventarios(wb.active, emp, arts, por_art, _periodo(None)); wb.active.title = "Libro de Inventarios"
    return _xlsx_response(wb, "libro_inventarios.xlsx")


@router.get("/catalogo.xlsx")
async def libro_catalogo(current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    cuentas = (await db.execute(select(CatalogoCuenta).where(
        CatalogoCuenta.empresa_id == emp.id, CatalogoCuenta.activa == True).order_by(CatalogoCuenta.codigo))).scalars().all()
    wb = Workbook(); _hoja_catalogo(wb.active, emp, cuentas); wb.active.title = "Catálogo de Cuentas"
    return _xlsx_response(wb, "catalogo_cuentas.xlsx")


@router.get("/balance.xlsx")
async def libro_balance(mes: int | None = Query(None, ge=1, le=12),
                        current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    emp = await _empresa(db, current_user.empresa_id)
    filas = await _lineas(db, emp.id, mes)
    wb = Workbook(); _hoja_balance(wb.active, emp, filas, _periodo(mes)); wb.active.title = "Balance Comprobación"
    return _xlsx_response(wb, f"balance_comprobacion_{mes or 'Q1'}.xlsx")


@router.get("/todos.xlsx")
async def libros_todos(mes: int | None = Query(None, ge=1, le=12),
                       current_user: Usuario = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Un solo archivo con todas las hojas: Diario, Mayor, Compras, Ventas, Inventarios, Catálogo, Balance."""
    emp = await _empresa(db, current_user.empresa_id)
    filas = await _lineas(db, emp.id, mes)
    cuentas = (await db.execute(select(CatalogoCuenta).where(
        CatalogoCuenta.empresa_id == emp.id, CatalogoCuenta.activa == True).order_by(CatalogoCuenta.codigo))).scalars().all()
    qc = select(LibroIvaCompra).where(LibroIvaCompra.empresa_id == emp.id)
    qv = select(LibroIvaVenta).where(LibroIvaVenta.empresa_id == emp.id)
    if mes:
        qc = qc.where(extract("month", LibroIvaCompra.fecha) == mes)
        qv = qv.where(extract("month", LibroIvaVenta.fecha) == mes)
    compras = (await db.execute(qc.order_by(LibroIvaCompra.fecha))).scalars().all()
    ventas = (await db.execute(qv.order_by(LibroIvaVenta.fecha))).scalars().all()
    arts = (await db.execute(select(ArticuloInventario).where(
        ArticuloInventario.empresa_id == emp.id).order_by(ArticuloInventario.codigo_sku))).scalars().all()
    movs = (await db.execute(select(MovimientoInventario).where(
        MovimientoInventario.empresa_id == emp.id).order_by(MovimientoInventario.fecha, MovimientoInventario.id))).scalars().all()
    por_art = defaultdict(list)
    for m in movs: por_art[m.articulo_id].append(m)
    per = _periodo(mes)

    wb = Workbook()
    _hoja_diario(wb.active, emp, filas, per); wb.active.title = "Libro Diario"
    _hoja_mayor(wb.create_sheet("Libro Mayor"), emp, filas, per, None)
    _hoja_compras(wb.create_sheet("Libro de Compras"), emp, compras, per)
    _hoja_ventas(wb.create_sheet("Libro de Ventas"), emp, ventas, per)
    _hoja_inventarios(wb.create_sheet("Libro de Inventarios"), emp, arts, por_art, per)
    _hoja_catalogo(wb.create_sheet("Catálogo"), emp, cuentas)
    _hoja_balance(wb.create_sheet("Balance Comprobación"), emp, filas, per)
    return _xlsx_response(wb, f"libros_legales_{mes or 'Q1'}_2025.xlsx")
